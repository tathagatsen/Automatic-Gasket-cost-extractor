import openpyxl
from fractions import Fraction

def getInch(inch):
    res = ""
    check = 0
    for i in inch:
        if i.isdigit():
            res += i
        if i == '.':
            check = 1
            res += i
        if i == '/':
            res += i
        if i.isalpha():
            break
    if check == 1:
        
        return Fraction(float(res))
    
    return Fraction(res)

def getRating(rate):
    res = []
    for char in rate:
        if char.isdigit():
            res.append(char)
    return int("".join(res))

# Load workbooks
db1 = openpyxl.load_workbook("BOQ 1.xlsx")
db2 = openpyxl.load_workbook("BOQ.xlsx", data_only=True)

# Select worksheets
ws1 = db1["Table 1"]
ws2 = db2["CS 0.5 TO 24"]

# Read data from ws1
Desc = [ws1[f"C{i}"].value for i in range(4, 19)]
Desc_M = [desc.split(',') for desc in Desc]

Inches = []
Rating = []
MOC = []
for i in range(len(Desc_M)):
    Inches.append(getInch(Desc_M[i][4]))
    Rating.append(getRating(Desc_M[i][5]))
    MOC.append(Desc_M[i][2].replace("L","-SS316/FG-SS316"))

# print(MOC)

# Read data from ws2
Rating_check = [ws2[f"A{i}"].value for i in range(47, 621)]
Inches_check = [ws2[f"B{i}"].value for i in range(47, 621)]

for i in range(len(Inches_check)):
    if Inches_check[i] == '' or Inches_check[i] is None:
        continue
    if isinstance(Inches_check[i], str):
        g = Inches_check[i].split()
        Inches_check[i] = float(g[0]) + float(Fraction(g[1]))

# Print final results (optional)
# print(Inches_check)

# print(len(Inches))
# print(len(MOC))
# print(len(Rating))
# print(len(Inches_check))
# print(len(Rating_check))

row1=ws1.max_row
row2 = ws2.max_row



# for i in range(len(Inches)):
#     for j in range(3, row2 + 1):
#         if Inches[i] == ws2[f"C{j}"].value and Rating[i] == ws2[f"A{j}"].value and (MOC[i]+"/FG-SS316" in ws2[f"D{j}"].value):
#             # Extract required values
#             P = ws2[f"P{j}"].value or 0
#             Q = ws2[f"Q{j}"].value or 0
#             O = ws2[f"O{j}"].value or 0
#             Y = ws2[f"Y{j}"].value or 0
#             W = ws2[f"W{j}"].value or 0

#             # Compute R3 and S3
#             R = P * Q
#             S = O * Q

#             # Compute T3 and U3
#             T = R * Q * 475
#             U = S * Q * 450

#             # Compute the final answer
#             ans = Y + W + U + T
#             print(ws2[f"A{j}"])
#             print(f"Inches: {Inches[i]}, Rating: {Rating[i]}, Computed Value: {ans}")
# print(MOC)
matching_val=[]
for i in range(len(Inches)):
    for j in range(3, row2 + 1):
        if Inches[i] == ws2[f"C{3+j}"].value and Rating[i] == ws2[f"A{3+j}"].value and (MOC[i] in ws2[f"D{3+j}"].value):
            # H = ws2[f"H{3+j}"].value or 0
            # G = ws2[f"G{3+j}"].value or 0
            # M = 0.014
            # Q = 1.000
            # Y = ws2[f"Y{3+j}"].value or 0
            # W = ws2[f"W{3+j}"].value or 0
            
            # J = round((H + G) / 2000,2)
            # K = round(((H - G) / 2) * 1.2)
            # L = K + 6
            # N = 0.024
            # O = round(J * K * M,3)
            # S = round(O * Q,3)
            # U = round(S * Q * 500,3)
            # P = round(J * L * N,3)
            # R = round(P * Q,3)
            # T = round(R * Q * 475,3)
            # ans = Y + W + U + T
            # print(G,H,J,K,M,N,O,P,Q,R,S,T,U,W,Y)
            
            ws1[f"I{4 + i}"].value=ws2[f"Z{3+j}"].value
            print(Inches[i], Rating[i], ws2[f"Z{3+j}"].value)
        # else:
        #     print(Inches[i], Rating[i], ws2[f"Z{3+j}"].value)
        if ws2[f"C{3+j}"].value==0.75 and  ws2[f"A{3+j}"].value==300 and (ws2[f"D{3+j}"].value=="SS316-SS316/FG-SS316"):
            print("hell")
db1.save("BOQ 1.xlsx")


