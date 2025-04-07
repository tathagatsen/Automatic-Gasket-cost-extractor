from flask import Flask, render_template, request, send_file, redirect, url_for, session
from werkzeug.utils import secure_filename
import openpyxl
import os
import io
from io import BytesIO
from openpyxl.utils import column_index_from_string
import tempfile
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = 'your_secret_key'

if os.path.exists('temp/ARC_gaskets.xlsx'):
    REFERENCE_WORKBOOK = openpyxl.load_workbook('temp/ARC_gaskets.xlsx')
else:
    print("Reference file not found!")


def process_excel(file1, file2,sheet1,sheet2, A1, A2, A3, A4, A5, A6, B1, B2, B3, B4, B5, B6, op, cost):
    def valid(col): return col != ""

    # Convert non-empty column letters to indices
    A_cols = [col for col in [A1, A2, A3, A4, A5, A6] if valid(col)]
    B_cols = [col for col in [B1, B2, B3, B4, B5, B6] if valid(col)]
    print(A_cols,B_cols)
    A_indices = [column_index_from_string(col) for col in A_cols]
    B_indices = [column_index_from_string(col) for col in B_cols]
    print(A_indices,B_indices)
    op = column_index_from_string(op)
    cost = column_index_from_string(cost)

    # Load both workbooks
    wb1 = load_workbook(filename=io.BytesIO(file1.read()))
    wb2 = load_workbook(filename=io.BytesIO(file2.read()))
    print(wb1,wb2)
    ws1 = wb1[sheet1]  # ENQUIRY
    ws2 = wb2[sheet2]  # COSTING
    print(ws1,ws2)
    row1 = ws1.max_row
    row2 = ws2.max_row

    # Build costing dictionary with dynamic key length
    costing_dict = {}
    for j in range(2, row2 + 1):
        key = tuple(ws2.cell(row=j, column=col).value for col in B_indices)
        costing_dict[key] = ws2.cell(row=j, column=cost).value

    # Match and update ENQUIRY sheet
    for i in range(2, row1 + 1):
        match_key = tuple(ws1.cell(row=i, column=col).value for col in A_indices)
        if match_key in costing_dict:
            ws1.cell(row=i, column=op).value = costing_dict[match_key]

    # Save modified workbook to memory
    output_stream = io.BytesIO()
    wb1.save(output_stream)
    output_stream.seek(0)
    return output_stream

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file1 = request.files['excel_file_1']
        file2 = request.files['excel_file_2']

        temp_file1 = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        file1.save(temp_file1.name)
        file1_path = temp_file1.name

        temp_file2 = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        file2.save(temp_file2.name)
        file2_path = temp_file2.name

        session['file1'] = file1_path
        session['file2'] = file2_path
        session['columns'] = {
            'A1': request.form['1A'].upper(),
            'B1': request.form['1B'].upper(),
            'A2': request.form['2A'].upper(),
            'B2': request.form['2B'].upper(),
            'A3': request.form['3A'].upper(),
            'B3': request.form['3B'].upper(),
            'A4': request.form['4A'].upper(),
            'B4': request.form['4B'].upper(),
            'A5': request.form['5A'].upper(),
            'B5': request.form['5B'].upper(),
            'A6': request.form['6A'].upper(),
            'B6': request.form['6B'].upper(),
            'op': request.form['op'].upper(),
            'cost': request.form['cost'].upper()
        }

        return redirect(url_for('home'))

    return render_template('index.html')

@app.route('/home', methods=['GET', 'POST'])
def home():
    if request.method == 'POST':
        sheet1 = request.form['sheet1']
        sheet2 = request.form['sheet2']
        columns = session.get('columns')
        file1_path = session.get('file1')
        file2_path = session.get('file2')
        if not file1_path or not file2_path:
            return redirect(url_for('index'))
        with open(file1_path, 'rb') as f1, open(file2_path, 'rb') as f2:
            modified_file = process_excel(
                f1, f2,sheet1,sheet2,
                columns['A1'], columns['A2'], columns['A3'],columns['A4'],columns['A5'], columns['A6'],columns['B1'], columns['B2'], columns['B3'],
                columns['B4'], columns['B5'], columns['B6'],
                columns['op'], columns['cost']
            )

        os.remove(file1_path)
        os.remove(file2_path)

        return send_file(modified_file, as_attachment=True, download_name='modified_file.xlsx')

    file1_path = session.get('file1')
    file2_path = session.get('file2')
    wb1 = openpyxl.load_workbook(file1_path)
    wb2 = openpyxl.load_workbook(file2_path)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    return render_template('home.html', sheets1=sheets1, sheets2=sheets2)

if __name__ == '__main__':
    app.run(debug=True)
